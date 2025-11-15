from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Avg, Q
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from .models import User, Quiz, Question, QuizAttempt, StudentAnswer
from .forms import StudentRegistrationForm, AdminRegistrationForm, LoginForm, QuizForm, QuestionForm
import random
import string
import openpyxl
from decouple import config  # For reading environment variables
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO


def generate_captcha(request):
    """Generate a random 6-character captcha with lowercase letters and digits"""
    import random
    import string
    captcha = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
    request.session['captcha'] = captcha
    return captcha


@csrf_protect
def register_view(request):
    print("Register view called")
    print("Request method:", request.method)
    
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Get form data
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        roll_number = request.POST.get('roll_number')
        branch = request.POST.get('branch')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        role = request.POST.get('role', 'student')
        
        # Force role to 'student' for all public registrations
        # Only superusers can create admin accounts
        role = 'student'
        
        # Validate form data
        errors = []
        if not username:
            errors.append('Username is required')
        if not email:
            errors.append('Email is required')
        if not phone:
            errors.append('Phone is required')
        if not roll_number:
            errors.append('Roll number is required')
        if not branch:
            errors.append('Branch is required')
        if not password1:
            errors.append('Password is required')
        if password1 != password2:
            errors.append('Passwords do not match')
        
        # Check if user already exists
        if User.objects.filter(username=username).exists():
            errors.append('Username already exists')
        if User.objects.filter(email=email).exists():
            errors.append('Email already exists')
        if User.objects.filter(phone=phone).exists():
            errors.append('Phone number already exists')
        
        if not errors:
            # Create user - only students can register publicly
            user = User.objects.create_user(
                username=username,
                email=email,
                phone=phone,
                roll_number=roll_number,
                branch=branch,
                password=password1,
                role='student'
            )
            
            messages.success(request, f'Registration successful! Welcome {user.username}. Please login to continue.')
            return redirect('login')
        else:
            # Add all errors as messages
            for error in errors:
                messages.error(request, error)
            return render(request, 'quiz/auth.html', {'show_register': True})
    else:
        print("GET request - showing registration form")
    
    return render(request, 'quiz/auth.html', {'show_register': True})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # Generate a new CAPTCHA only for GET requests
    if request.method == 'GET':
        captcha = generate_captcha(request)
    else:
        # For POST requests, use the existing CAPTCHA from session
        captcha = request.session.get('captcha', '')
    
    if request.method == 'POST':
        # Get username, password, and CAPTCHA from the POST data
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_captcha = request.POST.get('captcha')
        session_captcha = request.session.get('captcha')
        
        # Validate CAPTCHA
        if not user_captcha:
            messages.error(request, 'Please enter the CAPTCHA code.')
        elif user_captcha != session_captcha:
            messages.error(request, 'Invalid CAPTCHA code. Please try again.')
            # Generate a new CAPTCHA for the next attempt
            captcha = generate_captcha(request)
        elif username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {user.username}! Login successful.')
                return redirect('dashboard')
            else:
                messages.error(request, 'Login failed. Invalid username or password. Please try again.')
        else:
            messages.error(request, 'Login failed. Please fill in all required fields.')
    
    return render(request, 'quiz/auth.html', {'show_register': False, 'captcha': captcha})


@login_required
def logout_view(request):
    username = request.user.username
    logout(request)
    # Add a success message for logout
    messages.success(request, f'Goodbye, {username}! You have been successfully logged out.')
    return redirect('login')


@login_required
def dashboard_view(request):
    if request.user.role == 'admin':
        return redirect('admin_dashboard')
    else:
        return redirect('student_dashboard')


@login_required
def admin_dashboard(request):
    # Only allow superusers to access admin dashboard
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin dashboard is only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    # Show all quizzes for admin users, not just their own
    quizzes = Quiz.objects.all().annotate(
        attempt_count=Count('attempts')
    )
    
    active_quizzes_count = quizzes.filter(is_active=True).count()
    
    context = {
        'quizzes': quizzes,
        'total_quizzes': quizzes.count(),
        'total_attempts': QuizAttempt.objects.count(),
        'active_quizzes_count': active_quizzes_count,
    }
    
    return render(request, 'quiz/admin_dashboard.html', context)


@login_required
def student_dashboard(request):
    if request.user.role != 'student':
        messages.error(request, 'Access denied')
        return redirect('admin_dashboard')
    
    # Get all active quizzes
    all_quizzes = Quiz.objects.filter(is_active=True)
    
    # Get quizzes already attempted by the student
    attempted_quiz_ids = QuizAttempt.objects.filter(
        student=request.user, 
        is_completed=True
    ).values_list('quiz_id', flat=True)
    
    # Available quizzes (not attempted)
    available_quizzes = all_quizzes.exclude(id__in=attempted_quiz_ids)
    
    # Attempted quizzes with scores
    attempted_quizzes = QuizAttempt.objects.filter(
        student=request.user,
        is_completed=True
    ).select_related('quiz')
    
    context = {
        'available_quizzes': available_quizzes,
        'attempted_quizzes': attempted_quizzes,
    }
    
    return render(request, 'quiz/student_dashboard.html', context)


@login_required
def add_quiz(request):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        form = QuizForm(request.POST)
        if form.is_valid():
            quiz = form.save(commit=False)
            quiz.created_by = request.user
            quiz.save()
            messages.success(request, 'Quiz created successfully!')
            return redirect('add_questions', quiz_id=quiz.id)
    else:
        form = QuizForm()
    
    return render(request, 'quiz/add_quiz.html', {'form': form})


@login_required
def add_questions(request, quiz_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = Question.objects.filter(quiz=quiz)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.quiz = quiz
            question.save()
            messages.success(request, 'Question added successfully!')
            return redirect('add_questions', quiz_id=quiz.id)
    else:
        form = QuestionForm()
    
    context = {
        'quiz': quiz,
        'questions': questions,
        'form': form,
    }
    
    return render(request, 'quiz/add_questions.html', context)


@login_required
def delete_question(request, question_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    question = get_object_or_404(Question, id=question_id)
    quiz_id = question.quiz.id
    question.delete()
    messages.success(request, 'Question deleted successfully!')
    return redirect('add_questions', quiz_id=quiz_id)


@login_required
def edit_question(request, question_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    question = get_object_or_404(Question, id=question_id)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Question updated successfully!')
            return redirect('add_questions', quiz_id=question.quiz.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = QuestionForm(instance=question)
    
    return render(request, 'quiz/edit_question.html', {
        'form': form,
        'question': question
    })


@login_required
def toggle_quiz_status(request, quiz_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    quiz.is_active = not quiz.is_active
    quiz.save()
    
    status = 'activated' if quiz.is_active else 'deactivated'
    messages.success(request, f'Quiz "{quiz.title}" has been {status} successfully!')
    return redirect('admin_dashboard')


@login_required
def delete_quiz(request, quiz_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    if request.method == 'POST':
        quiz_title = quiz.title
        quiz.delete()
        messages.success(request, f'Quiz "{quiz_title}" has been deleted successfully!')
        return JsonResponse({'status': 'success', 'message': f'Quiz "{quiz_title}" has been deleted successfully!'})
    
    # For GET requests, return JSON response for AJAX modal
    return JsonResponse({'status': 'confirm', 'quiz_id': quiz_id, 'quiz_title': quiz.title})


@login_required
def view_results(request):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quizzes = Quiz.objects.all()
    selected_quiz_id = request.GET.get('quiz_id')
    
    attempts = None
    selected_quiz = None
    
    if selected_quiz_id:
        selected_quiz = get_object_or_404(Quiz, id=selected_quiz_id)
        attempts = QuizAttempt.objects.filter(
            quiz=selected_quiz,
            is_completed=True
        ).select_related('student').order_by('student__roll_number')
    
    context = {
        'quizzes': quizzes,
        'selected_quiz': selected_quiz,
        'attempts': attempts,
    }
    
    return render(request, 'quiz/view_results.html', context)


@login_required
def export_results_excel(request, quiz_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    attempts = QuizAttempt.objects.filter(
        quiz=quiz,
        is_completed=True
    ).select_related('student').order_by('-score')
    
    # Create workbook with enhanced styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Results"
    
    # Enhanced styling
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Color palette
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")  # Blue
    header_font = Font(bold=True, color="FFFFFF", size=12)
    alternate_fill1 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Light gray
    alternate_fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Add title with enhanced styling
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f'Quiz Results: {quiz.title}'
    title_cell.font = Font(bold=True, size=16, color="1E40AF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    # Add subtitle with date
    from datetime import datetime
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
    subtitle_cell.font = Font(italic=True, size=10, color="6B7280")
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers with enhanced styling
    headers = ['S.No', 'Student Name', 'Roll Number', 'Score', 'Total Marks', 'Percentage']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data with alternating row colors and enhanced styling
    for idx, attempt in enumerate(attempts, 1):
        row_num = idx + 4
        # Alternate row colors
        row_fill = alternate_fill1 if idx % 2 == 0 else alternate_fill2
        
        # S.No
        cell = ws.cell(row=row_num, column=1, value=idx)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Student Name
        cell = ws.cell(row=row_num, column=2, value=attempt.student.username)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Roll Number
        cell = ws.cell(row=row_num, column=3, value=attempt.student.roll_number or 'N/A')
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Score
        cell = ws.cell(row=row_num, column=4, value=attempt.score)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        # Highlight high scores
        if attempt.percentage() >= 75:
            cell.font = Font(color="16A34A", bold=True)  # Green for excellent
        elif attempt.percentage() >= 50:
            cell.font = Font(color="CA8A04")  # Yellow for good
        else:
            cell.font = Font(color="DC2626")  # Red for poor
        
        # Total Marks
        cell = ws.cell(row=row_num, column=5, value=attempt.total_marks)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Percentage with color coding
        cell = ws.cell(row=row_num, column=6, value=f"{attempt.percentage()}%")
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        # Color code percentages
        if attempt.percentage() >= 75:
            cell.font = Font(color="16A34A", bold=True)  # Green for excellent
        elif attempt.percentage() >= 50:
            cell.font = Font(color="CA8A04")  # Yellow for good
        else:
            cell.font = Font(color="DC2626")  # Red for poor
    
    # Adjust column widths for better readability
    column_widths = {
        1: 8,   # S.No
        2: 25,  # Student Name
        3: 15,  # Roll Number
        4: 10,  # Score
        5: 12,  # Total Marks
        6: 12   # Percentage
    }
    
    for col_idx, width in column_widths.items():
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_results_pdf(request, quiz_id):
    # Only allow superusers to access admin features
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin features are only accessible to superusers.')
        return redirect('student_dashboard')
    
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('student_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # Order by roll number ascending
    attempts = QuizAttempt.objects.filter(
        quiz=quiz,
        is_completed=True
    ).select_related('student').order_by('student__roll_number')
    
    # Create PDF with enhanced styling
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
    elements = []
    
    # Enhanced styles
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        fontSize=24,
        textColor=HexColor('#1E3A8A'),
        spaceAfter=25,
        alignment=1,  # Center
        fontName='Helvetica-Bold'
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        fontSize=14,
        textColor=HexColor('#4B5563'),
        spaceAfter=35,
        alignment=1,  # Center
        fontName='Helvetica'
    )
    
    # Header style
    header_style = ParagraphStyle(
        'Header',
        fontSize=16,
        textColor=HexColor('#FFFFFF'),
        spaceAfter=20,
        alignment=1,  # Center
        fontName='Helvetica-Bold'
    )
    
    # Add title and subtitle
    from datetime import datetime
    from reportlab.platypus import Paragraph
    
    title = Paragraph(f"Quiz Results: {quiz.title}", title_style)
    subtitle = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style)
    
    elements.append(title)
    elements.append(subtitle)
    elements.append(Spacer(1, 30))
    
    # Create table data with only roll number, email, and score
    data = [['S.No', 'Roll Number', 'Email', 'Score']]
    
    # Add data rows - show only clean score value
    for idx, attempt in enumerate(attempts, 1):
        data.append([
            str(idx),
            attempt.student.roll_number or 'N/A',
            attempt.student.email,
            str(attempt.score)  # Show only the numeric score value
        ])
    
    # Create table with enhanced styling
    table = Table(data, colWidths=[0.8*inch, 1.8*inch, 2.5*inch, 1.2*inch])
    
    # Table styling with attractive design
    table_style = TableStyle([
        # Header styling with gradient-like effect
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
        ('TOPPADDING', (0, 0), (-1, 0), 15),
        
        # Data rows styling
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),    # S.No
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),    # Roll Number
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),      # Email
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),    # Score
        
        # Font styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (2, -1), 12),
        
        # Grid and borders with attractive styling
        ('GRID', (0, 0), (-1, -1), 2, HexColor('#1E40AF')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#EFF6FF'), HexColor('#FFFFFF')]),
        
        # Add some spacing
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Add summary statistics if there are attempts
    if attempts:
        elements.append(Spacer(1, 40))
        
        # Summary title
        summary_title = Paragraph("Performance Summary", header_style)
        elements.append(summary_title)
        elements.append(Spacer(1, 20))
        
        # Calculate statistics
        total_students = len(attempts)
        avg_score = sum([attempt.score for attempt in attempts]) / total_students if total_students > 0 else 0
        avg_percentage = sum([attempt.percentage() for attempt in attempts]) / total_students if total_students > 0 else 0
        highest_score = max([attempt.score for attempt in attempts]) if attempts else 0
        lowest_score = min([attempt.score for attempt in attempts]) if attempts else 0
        
        # Summary data with attractive styling
        summary_data = [
            ['Metric', 'Value'],
            ['Total Students', str(total_students)],
            ['Average Score', f"{avg_score:.2f}"],
            ['Average Percentage', f"{avg_percentage:.2f}%"],
            ['Highest Score', str(highest_score)],
            ['Lowest Score', str(lowest_score)]
        ]
        
        # Create summary table
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 2, HexColor('#1E40AF')),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('BACKGROUND', (0, 1), (-1, -1), HexColor('#EFF6FF')),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.pdf"'
    
    return response


@login_required
def take_quiz(request, quiz_id):
    if request.user.role != 'student':
        messages.error(request, 'Access denied')
        return redirect('admin_dashboard')
    
    quiz = get_object_or_404(Quiz, id=quiz_id, is_active=True)
    
    # Check if student has already attempted this quiz
    existing_attempt = QuizAttempt.objects.filter(student=request.user, quiz=quiz).first()
    if existing_attempt and existing_attempt.is_completed:
        messages.error(request, 'You have already attempted this quiz')
        return redirect('student_dashboard')
    
    # Create or get attempt
    if not existing_attempt:
        questions = Question.objects.filter(quiz=quiz)
        total_marks = sum([q.marks for q in questions])
        
        attempt = QuizAttempt.objects.create(
            student=request.user,
            quiz=quiz,
            total_marks=total_marks
        )
    else:
        attempt = existing_attempt
    
    questions = Question.objects.filter(quiz=quiz).order_by('order', 'id')
    
    context = {
        'quiz': quiz,
        'questions': questions,
        'attempt': attempt,
    }
    
    return render(request, 'quiz/take_quiz.html', context)


@login_required
def submit_quiz(request, attempt_id):
    if request.user.role != 'student':
        messages.error(request, 'Access denied')
        return redirect('admin_dashboard')
    
    attempt = get_object_or_404(QuizAttempt, id=attempt_id, student=request.user)
    
    if attempt.is_completed:
        messages.error(request, 'This quiz has already been submitted')
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        questions = Question.objects.filter(quiz=attempt.quiz)
        score = 0
        
        for question in questions:
            answer_key = f'question_{question.id}'
            selected_answer = request.POST.get(answer_key)
            
            # Handle unanswered questions (when selected_answer is None or empty)
            if not selected_answer:
                selected_answer = None
            
            is_correct = False
            if selected_answer and selected_answer == question.correct_answer:
                is_correct = True
                score += question.marks
            
            StudentAnswer.objects.create(
                attempt=attempt,
                question=question,
                selected_answer=selected_answer,
                is_correct=is_correct
            )
        
        attempt.score = score
        attempt.is_completed = True
        attempt.completed_at = timezone.now()
        attempt.save()
        
        messages.success(request, f'Quiz submitted successfully! Your score: {score}/{attempt.total_marks}')
        return redirect('quiz_result', attempt_id=attempt.id)
    
    return redirect('take_quiz', quiz_id=attempt.quiz.id)


@login_required
def quiz_result(request, attempt_id):
    if request.user.role != 'student':
        messages.error(request, 'Access denied')
        return redirect('admin_dashboard')
    
    attempt = get_object_or_404(QuizAttempt, id=attempt_id, student=request.user)
    
    if not attempt.is_completed:
        messages.error(request, 'Quiz not yet completed')
        return redirect('take_quiz', quiz_id=attempt.quiz.id)
    
    answers = StudentAnswer.objects.filter(attempt=attempt).select_related('question')
    
    # Calculate correct, wrong, and unanswered answers
    correct_count = answers.filter(is_correct=True).count()
    wrong_count = answers.filter(is_correct=False, selected_answer__isnull=False).count()
    unanswered_count = answers.filter(selected_answer__isnull=True).count()
    total_questions = answers.count()
    
    context = {
        'attempt': attempt,
        'answers': answers,
        'correct_count': correct_count,
        'wrong_count': wrong_count,
        'unanswered_count': unanswered_count,
        'total_questions': total_questions,
    }
    
    return render(request, 'quiz/quiz_result.html', context)


@login_required
def profile_view(request):
    return render(request, 'quiz/profile.html', {'user': request.user})
